"""Parse QIMA payroll Excel workbooks into staging-ready row lists."""
from __future__ import annotations

import re
from io import BytesIO

import openpyxl

from src.payroll.columns import COLUMN_COUNT, EXPECTED_HEADERS


class ParseError(Exception):
    """Structural error in the workbook (wrong sheet, wrong columns, etc.)."""


def _normalise(text: str | None) -> str:
    """Collapse all whitespace (including embedded newlines) to single spaces, strip."""
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def _assert_headers(row2_values: tuple) -> None:
    """Assert row-2 headers match the frozen expected list (after normalisation)."""
    actual = [_normalise(v) for v in row2_values[:COLUMN_COUNT]]
    expected = [_normalise(h) for h in EXPECTED_HEADERS]

    if len(actual) < COLUMN_COUNT:
        raise ParseError(
            f"Expected {COLUMN_COUNT} columns, got {len(actual)}. "
            "Template structure may have changed."
        )

    mismatches = []
    for i, (a, e) in enumerate(zip(actual, expected)):
        if a != e:
            col_letter = openpyxl.utils.get_column_letter(i + 1)
            mismatches.append(f"  {col_letter} (pos {i+1}): got {a!r}, expected {e!r}")

    if mismatches:
        raise ParseError(
            f"Header mismatch at {len(mismatches)} column(s):\n"
            + "\n".join(mismatches[:10])
        )


def subsidiary_code(row: list) -> str:

    value = row[4]
    if not value:
        raise ParseError(
            "Column E (Subsidiary) is empty on a row with an Employee SAP ID. "
            "Cannot derive SUBSIDIARY_CODE."
        )
    parts = str(value).split(" - ", 1)
    return parts[0].strip()


def parse(file_bytes: bytes, sheet_name: str = "2026") -> list[list[str | None]]:

    wb = openpyxl.load_workbook(
        BytesIO(file_bytes), read_only=True, data_only=True
    )

    try:
        if sheet_name not in wb.sheetnames:
            raise ParseError(
                f"Sheet '{sheet_name}' not found. "
                f"Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Read row 2 (headers) and assert structure
        header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if header_row is None:
            raise ParseError("Row 2 (headers) is empty or missing.")

        _assert_headers(header_row)

        # Read data rows starting at row 3
        rows: list[list[str | None]] = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            # A data row must have EMPLOYEE_SAP_ID (column A, index 0)
            if row[0] is None:
                continue

            # Convert all 67 cells to str or None (rows can come back short of 67)
            rows.append([
                str(row[i]) if i < len(row) and row[i] is not None else None
                for i in range(COLUMN_COUNT)
            ])

        return rows

    finally:
        wb.close()
